from pathlib import Path
import json
from openpyxl import load_workbook
import itertools


def detect_header_from_rows(rows, required):
    """
    Bestimmt Header-Index (0-basiert) und Spaltenmapping (0-basiert) aus bis zu 10 vorab gelesenen Zeilen.
    """
    for ridx, row in enumerate(rows):
        values = [(str(v).strip() if v is not None else "") for v in row]
        header_map = {}
        for idx, val in enumerate(values):
            for key in required:
                if val.lower() == key.lower():
                    header_map[key] = idx
        if all(k in header_map for k in required):
            return ridx, header_map
    # Fallback: erste Zeile als Header in Standardreihenfolge interpretieren
    return 0, {k: i for i, k in enumerate(required)}


def coerce_int(value):
    try:
        return int(value)
    except Exception:
        try:
            return int(float(value))
        except Exception:
            return None


def coerce_float(value):
    if value is None:
        return None
    s = str(value).strip().replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def main():
    root = Path(__file__).resolve().parents[1]
    excel_update = root / "Ergebnisse_update.xlsx"
    excel_orig = root / "Ergebnisse.xlsx"
    excel_path = excel_update if excel_update.exists() else excel_orig
    if not excel_path.exists():
        raise SystemExit("Excel-Datei nicht gefunden (Ergebnisse_update.xlsx / Ergebnisse.xlsx).")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    required = ["Platz", "Jahr", "Band", "Album", "Note"]
    # Stream-Zugriff: zuerst bis zu 10 Zeilen puffern, um Header zu erkennen
    row_iter = ws.iter_rows(values_only=True)
    first_rows = list(itertools.islice(row_iter, 10))
    if not first_rows:
        raise SystemExit("Arbeitsblatt ist leer.")
    header_row_idx, header_map = detect_header_from_rows(first_rows, required)

    records = []
    # Kette: verbleibende gepufferte Zeilen nach Header + restlicher Stream
    remaining_buffered = first_rows[header_row_idx + 1 :]
    data_iter = itertools.chain(remaining_buffered, row_iter)
    # Header-Indices sind 0-basiert
    idx_pl = header_map["Platz"]
    idx_j = header_map["Jahr"]
    idx_b = header_map["Band"]
    idx_a = header_map["Album"]
    idx_n = header_map["Note"]

    for row in data_iter:
        # Zeilen können unterschiedlich lang sein → defensiver Zugriff
        def get_idx(row_vals, i):
            return row_vals[i] if i < len(row_vals) else None

        platz = get_idx(row, idx_pl)
        jahr = get_idx(row, idx_j)
        band = get_idx(row, idx_b)
        album = get_idx(row, idx_a)
        note = get_idx(row, idx_n)

        platz_i = coerce_int(platz)
        jahr_i = coerce_int(jahr)
        band_s = (str(band).strip() if band is not None else "")
        album_s = (str(album).strip() if album is not None else "")
        note_f = coerce_float(note)

        if jahr_i is None or band_s == "" or album_s == "":
            continue

        records.append(
            {
                "Platz": platz_i,
                "Jahr": jahr_i,
                "Band": band_s,
                "Album": album_s,
                "Note": note_f,
            }
        )

    out_dir = root / "public" / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "alben.json"
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False)

    print(f"Exportiert: {out_path} ({len(records)} Zeilen)")


if __name__ == "__main__":
    main()
